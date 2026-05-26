#!/usr/bin/env python3
"""
Genera Archivo Cronológico HTML — Desguaces VFU — Alvale Consulting Ingenieros, S.L.
"""

import sys, os, re, webbrowser, threading
from datetime import date, datetime

def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)

LOGO_DEFAULT = resource_path('logo_alvale.png')

# ── Utilidades ────────────────────────────────────────────────────
def _str(v): return '' if v is None else str(v).strip()
def _ler(v):
    if v is None: return ''
    try: return str(int(float(str(v))))
    except: return str(v).strip()
def _nima(v):
    if v is None: return ''
    try: return str(int(float(str(v))))
    except: return str(v).strip()
def _fecha(v):
    if v is None: return ''
    if isinstance(v, (datetime, date)): return v.strftime('%d/%m/%Y')
    s = str(v).strip()
    for fmt in ('%d/%m/%Y','%Y-%m-%d','%d-%m-%Y'):
        try: return datetime.strptime(s,fmt).strftime('%d/%m/%Y')
        except: pass
    return s
def _mun(v):
    if v is None: return ''
    s = re.sub(r'^\d+[-–]\s*','',str(v).strip()).strip()
    m = re.match(r'^(.+ .+?)-([A-ZÁÉÍÓÚÜÑ])',s)
    if m and len(m.group(1))>6: s=m.group(1).strip()
    return s.split('/')[0].strip() if '/' in s else s
def _titulo(texto):
    if not texto: return texto
    siglas=set(re.findall(r'\b(?:[A-ZÁÉÍÓÚÜÑ]\.){2,}',texto))
    r=texto.title()
    for s in siglas: r=r.replace(s.title(),s)
    return r
def _hoy_es():
    meses=['enero','febrero','marzo','abril','mayo','junio','julio','agosto','septiembre','octubre','noviembre','diciembre']
    h=date.today(); return f"{h.day} de {meses[h.month-1]} de {h.year}"

# ── Abreviaciones ─────────────────────────────────────────────────
PROV_ABBR={"Álava":"ARA","Araba":"ARA","Araba/Álava":"ARA","Albacete":"ALB","Alicante":"ALI","Alacant":"ALI","Almería":"ALM","Asturias":"AST","Ávila":"AVI","Badajoz":"BAD","Barcelona":"BCN","Bizkaia":"BIZ","Vizcaya":"BIZ","Burgos":"BUR","Cáceres":"CAC","Cádiz":"CAD","Cantabria":"CTB","Castellón":"CAS","Castelló":"CAS","Ciudad Real":"CRE","Córdoba":"COR","A Coruña":"ACO","Cuenca":"CUE","Girona":"GIR","Gerona":"GIR","Granada":"GRA","Guadalajara":"GUA","Gipuzkoa":"GIP","Guipúzcoa":"GIP","Huelva":"HUE","Huesca":"HUS","Illes Balears":"BAL","Baleares":"BAL","Jaén":"JAE","León":"LEO","Lleida":"LLE","Lérida":"LLE","Lugo":"LUG","Madrid":"MAD","Málaga":"MAL","Murcia":"MUR","Navarra":"NAV","Nafarroa":"NAV","Ourense":"OUR","Orense":"OUR","Palencia":"PAL","Las Palmas":"LPA","Pontevedra":"PON","La Rioja":"RIO","Rioja":"RIO","Salamanca":"SAL","Santa Cruz de Tenerife":"TFE","Segovia":"SEG","Sevilla":"SEV","Soria":"SOR","Tarragona":"TAR","Teruel":"TER","Toledo":"TOL","Valencia":"VAL","València":"VAL","Valladolid":"VLD","Zamora":"ZAM","Zaragoza":"ZGZ","Ceuta":"CEU","Melilla":"MEL"}
CCAA_ABBR={"Andalucía":"AND","Andalucia":"AND","Aragón":"ARA","Aragon":"ARA","Asturias":"AST","Principado de Asturias":"AST","Illes Balears":"BAL","Baleares":"BAL","Canarias":"CAN","Cantabria":"CTB","Castilla-La Mancha":"CLM","Castilla La Mancha":"CLM","Castilla y León":"CYL","Castilla y Leon":"CYL","Cataluña":"CAT","Catalunya":"CAT","Comunitat Valenciana":"CVA","Comunidad Valenciana":"CVA","Extremadura":"EXT","Galicia":"GAL","La Rioja":"RIO","Rioja":"RIO","Madrid":"MAD","Comunidad de Madrid":"MAD","Murcia":"MUR","Región de Murcia":"MUR","Navarra":"NAV","Comunidad Foral de Navarra":"NAV","Nafarroa":"NAV","País Vasco":"PV","Euskadi":"PV","Pais Vasco":"PV","Ceuta":"CEU","Melilla":"MEL"}
PAIS_ABBR={"España":"ES","Espana":"ES","Spain":"ES","Portugal":"PT","Francia":"FR","France":"FR","Alemania":"DE","Germany":"DE","Italia":"IT","Italy":"IT","Reino Unido":"UK","United Kingdom":"UK","Países Bajos":"NL","Holanda":"NL","Netherlands":"NL","Bélgica":"BE","Belgium":"BE","Suiza":"CH","Switzerland":"CH","Austria":"AT","Polonia":"PL","Poland":"PL","Estados Unidos":"US","USA":"US","Marruecos":"MA","Morocco":"MA"}

def _abbr(v, table):
    if not v: return ''
    s=re.sub(r'^\d+\s*[-–]\s*','',str(v).strip())
    if s in table: return table[s]
    sl=s.lower()
    for k,a in table.items():
        if k.lower()==sl: return a
    for k,a in table.items():
        if sl in k.lower(): return a
    return s[:6]

# ── Lectura Excel ─────────────────────────────────────────────────
def leer_datos_centro(wb):
    centro={}
    try:
        ws=wb['DATOS EMPRESA']
        for row in ws.iter_rows(min_row=1,max_row=30,values_only=True):
            if not row[0] or row[1] is None: continue
            c=str(row[0]).strip().lower(); v=row[1]
            if 'razón social' in c or 'razon social' in c: centro['razon_social']=_str(v)
            elif c=='nif': centro['nif']=_str(v)
            elif c=='nima': centro['nima']=_nima(v)
            elif 'tipo' in c and ('autorización' in c or 'autorizacion' in c): centro['tipo_aut']=_str(v)
            elif 'autorización' in c or 'autorizacion' in c: centro['num_aut']=_str(v)
            elif 'dirección' in c or 'direccion' in c: centro['direccion']=_str(v)
            elif 'municipio' in c: centro['municipio']=_str(v)
            elif 'provincia' in c: centro['provincia']=_str(v)
            elif 'año' in c or 'ano' in c: centro['año']=_str(v)
    except KeyError: pass
    for k in ('razon_social','nif','nima','num_aut','tipo_aut','año'): centro.setdefault(k,'')
    return centro

def leer_salidas(wb):
    rows=[]
    try: ws=wb['REGISTRO SALIDAS']
    except KeyError: return rows
    for row in ws.iter_rows(min_row=3,values_only=True):
        if not (row[0] if len(row)>0 else None): continue
        def cv(i): return row[i] if len(row)>i else None
        try: peso_t=float(cv(4)) if cv(4) and not str(cv(4)).startswith('=') else float(cv(3))/1000
        except: peso_t=0.0
        rows.append({'fecha':_fecha(cv(0)),'ler':_ler(cv(1)),'denominacion':_titulo(_str(cv(2))),'cant_t':peso_t,'di':_str(cv(5)),'transp_insc':_str(cv(19)),'dest_razon':_titulo(_str(cv(7))),'dest_nif':_str(cv(8)),'dest_nima':_nima(cv(9)),'dest_aut':_str(cv(10)),'dest_tipo':_str(cv(11)),'municipio':_mun(cv(12)),'prov':_abbr(cv(13),PROV_ABBR),'ccaa':_abbr(cv(14),CCAA_ABBR),'pais':_abbr(cv(15),PAIS_ABBR),'transp_razon':_titulo(_str(cv(16))),'metodo':_str(cv(20)),'proceso':_str(cv(21)).replace('Desmontaje VFU','Desmontaje').replace('desmontaje VFU','Desmontaje')})
    return rows

# ── HTML ──────────────────────────────────────────────────────────
def _e(t):
    if not t: return ''
    return str(t).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')
def _css_str(s): return str(s).replace('\\','\\\\').replace('"','\\"')
def _fmt_t(v):
    try:
        f=float(v); s=f'{f:.3f}'.replace('.',','); p=s.split(',')
        i='';
        for n,ch in enumerate(reversed(p[0])):
            if n>0 and n%3==0: i='.'+i
            i=ch+i
        return i+','+(p[1] if len(p)>1 else '000')
    except: return str(v)
def _logo_b64(path):
    if not os.path.exists(path): return None
    import base64
    with open(path,'rb') as f: return 'data:image/png;base64,'+base64.b64encode(f.read()).decode()

def _cover_row(label,value,bold=False,last=False):
    b='' if last else 'border-bottom:0.2mm solid #E8E8EF;'
    w='font-weight:700;' if bold else 'font-weight:600;'
    return f'<tr><td style="padding:2.6mm 4mm;{b}font-size:9.5pt;color:#2C2E88;font-weight:600;width:44mm;">{label}</td><td style="padding:2.6mm 4mm;{b}font-size:10pt;color:#0B0C16;{w}">{value}</td></tr>'

def _colgroup():
    return '<colgroup><col style="width:5.8%"/><col style="width:3.2%"/><col style="width:7%"/><col style="width:3.5%"/><col style="width:10%"/><col style="width:9.5%"/><col style="width:4.5%"/><col style="width:5.5%"/><col style="width:6%"/><col style="width:2.8%"/><col style="width:4.5%"/><col style="width:2.8%"/><col style="width:2.5%"/><col style="width:2.2%"/><col style="width:9.5%"/><col style="width:6%"/><col style="width:4.7%"/><col style="width:6%"/></colgroup>'

def _thead():
    return '<thead><tr class="group"><th>Fecha</th><th colspan="2">Residuo</th><th>Cant.</th><th style="white-space:nowrap;">Documentación</th><th colspan="5">Destino</th><th colspan="4">Ubic. destino</th><th colspan="2">Transportista</th><th>Trat.</th><th>Proceso</th></tr><tr class="sub"><th>Fecha<br/>traslado</th><th>Cód.<br/>LER</th><th>Descripción</th><th class="num">(t)</th><th>Nº DI</th><th class="th-razon">Razón social</th><th>NIF</th><th>NIMA</th><th>Nº aut.</th><th>Tipo<br/>insc.</th><th>Mun.</th><th>Prov.</th><th>CCAA</th><th>País</th><th class="th-razon">Razón social</th><th>Nº insc.<br/>transp.</th><th>Método<br/>valor.</th><th>Proceso</th></tr></thead>'

def _row(r):
    return f'<tr><td class="date">{_e(r["fecha"])}</td><td class="ler">{_e(r["ler"])}</td><td>{_e(r["denominacion"])}</td><td class="num">{_fmt_t(r["cant_t"])}</td><td class="mono-long">{_e(r["di"])}</td><td>{_e(r["dest_razon"])}</td><td class="mono-long">{_e(r["dest_nif"])}</td><td class="mono">{_e(r["dest_nima"])}</td><td class="mono-long">{_e(r["dest_aut"])}</td><td class="abbr">{_e(r["dest_tipo"])}</td><td>{_e(r["municipio"])}</td><td class="abbr">{_e(r["prov"])}</td><td class="abbr">{_e(r["ccaa"])}</td><td class="abbr">{_e(r["pais"])}</td><td>{_e(r["transp_razon"])}</td><td class="mono-long">{_e(r["transp_insc"])}</td><td class="valor">{_e(r["metodo"]).replace("/", "/<br/>")}</td><td class="valor">{_e(r["proceso"])}</td></tr>'

def _get_css(nombre='',num_aut='',periodo=''):
    return """
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:ital,wght@0,400;0,500;0,600;0,700;1,400&family=IBM+Plex+Mono:wght@400;500;600&family=Poppins:wght@400;600;700&display=swap');
:root{--brand-navy:#2C2E88;--brand-green:#238D3E;--brand-purple:#5D549D;--brand-lavender:#8B83BA;--navy-50:#EEEFF7;--navy-700:#1B1D58;--ink-25:#FAFAFC;--ink-50:#F4F4F8;--ink-100:#E8E8EF;--ink-200:#D2D2DE;--ink-400:#7E8094;--ink-700:#2A2C3D;--font-body:'IBM Plex Sans','Aptos',system-ui,sans-serif;--font-mono:'IBM Plex Mono','Cascadia Mono','Consolas',monospace;--font-brand:'Poppins','IBM Plex Sans',system-ui,sans-serif;}
*{box-sizing:border-box;}h1,h2,h3{margin:0;padding:0;font-size:inherit;font-weight:inherit;}
.pdf-bookmark{display:inline;font-size:inherit;font-weight:inherit;color:inherit;font-family:inherit;letter-spacing:inherit;}
html,body{margin:0;padding:0;font-family:var(--font-body);color:var(--ink-700);background:#d9dae0;-webkit-font-smoothing:antialiased;}
.sheet--cover{width:210mm;height:297mm;background:#fff;margin:18px auto;box-shadow:0 2px 10px rgba(0,0,0,.10);page-break-after:always;display:flex;flex-direction:column;overflow:hidden;}
.cover__eyebrow{display:flex;gap:3px;margin-bottom:7mm;}
.cover__eyebrow span{display:block;width:20px;height:10px;border-radius:1px;}
.cover__eyebrow span:nth-child(1){background:var(--brand-navy);}
.cover__eyebrow span:nth-child(2){background:var(--brand-green);}
.cover__eyebrow span:nth-child(3){background:var(--brand-purple);}
.cover__eyebrow span:nth-child(4){background:var(--brand-lavender);}
.cover__eyebrow span:nth-child(5){background:#C3C0DE;}
.crono-section{background:#fff;max-width:297mm;margin:18px auto;padding:6mm 10mm 10mm;box-shadow:0 2px 10px rgba(0,0,0,.10);}
.crono-section .doc-section{margin-bottom:2mm;}
.doc-section__title{font-family:var(--font-body);font-weight:700;font-size:10pt;color:var(--brand-navy);margin:0 0 1.5mm;line-height:1.1;text-transform:uppercase;letter-spacing:0.06em;}
.crono-table{width:100%;border-collapse:collapse;font-size:5.8pt;font-family:var(--font-body);table-layout:fixed;line-height:1.18;color:var(--ink-700);}
.crono-table thead .group th{background:var(--brand-navy);color:#fff;font-weight:800;letter-spacing:0.06em;text-transform:uppercase;font-size:6.6pt;padding:1.1mm 0.8mm;text-align:center;border-right:0.25mm solid rgba(255,255,255,.35);border-bottom:0.25mm solid rgba(255,255,255,.5);vertical-align:middle;white-space:nowrap;}
.crono-table thead .group th:last-child{border-right:0;}
.crono-table thead .sub th{background:#ADB1D9;color:#1B1D58;font-weight:700;font-size:6pt;padding:1mm 0.8mm;text-align:center;border-right:0.15mm solid rgba(255,255,255,.18);vertical-align:middle;line-height:1.15;}
.crono-table thead .sub th:last-child{border-right:0;}
.crono-table thead{display:table-header-group;}
.crono-table tbody td{padding:0.9mm 0.7mm;border-bottom:0.15mm solid var(--ink-100);border-right:0.1mm solid var(--ink-50);vertical-align:middle;text-align:center;word-wrap:break-word;overflow-wrap:break-word;hyphens:auto;-webkit-hyphens:auto;font-size:5.8pt;line-height:1.18;}
.crono-table tbody tr:nth-child(even){background:var(--ink-25);}
.crono-table tbody td.num{font-family:var(--font-mono);font-variant-numeric:tabular-nums;white-space:nowrap;}
.crono-table tbody td.ler{font-family:var(--font-mono);white-space:nowrap;}
.crono-table tbody td.date{white-space:nowrap;font-variant-numeric:tabular-nums;font-family:var(--font-mono);}
.crono-table tbody td.mono{font-family:var(--font-mono);white-space:nowrap;}
.crono-table tbody td.mono-long{font-family:var(--font-mono);white-space:normal;word-break:break-all;overflow-wrap:anywhere;}
.crono-table tbody td.abbr{font-family:var(--font-mono);white-space:nowrap;font-weight:600;}
.crono-table tbody td.valor{white-space:normal;word-break:keep-all;line-height:1.15;}
.crono-table tr{break-inside:avoid;page-break-inside:avoid;}
.th-razon{background:#ADB1D9;color:#1B1D58;}
.print-bar{position:fixed;top:16px;right:16px;z-index:10;background:#fff;border-radius:8px;box-shadow:0 6px 16px -6px rgba(11,12,36,0.12);padding:6px;}
.print-bar button{font-family:var(--font-body);font-weight:600;font-size:13px;padding:8px 14px;border-radius:5px;border:none;background:var(--brand-navy);color:#fff;cursor:pointer;}
.print-bar button:hover{background:var(--navy-700);}
@media print{html,body{background:#fff;}.sheet--cover{margin:0;box-shadow:none;}.crono-section{padding:0;margin:0;box-shadow:none;max-width:none;}.print-bar{display:none !important;}}
""" + f"""
@page{{size:A4 landscape;margin:14mm 10mm 12mm 10mm;
  @top-left{{content:"{_css_str(nombre)} | {_css_str(num_aut)} | {_css_str(periodo)}";font-family:system-ui,sans-serif;font-size:7.5pt;font-weight:600;color:#2C2E88;vertical-align:bottom;padding-bottom:2mm;}}
  @top-right{{content:"Alvale Consulting Ingenieros, S.L.";font-family:system-ui,sans-serif;font-size:7.5pt;font-weight:600;color:#2C2E88;vertical-align:bottom;padding-bottom:2mm;text-align:right;}}
  @bottom-left{{content:"Archivo Cronológico — {_css_str(nombre)} — {_css_str(periodo)}";font-family:system-ui,sans-serif;font-size:7pt;color:#6B6D82;vertical-align:top;padding-top:2mm;}}
  @bottom-right{{content:"Página " counter(page) " de " counter(pages);font-family:system-ui,sans-serif;font-size:7pt;color:#6B6D82;vertical-align:top;padding-top:2mm;}}
}}
@page :first{{size:A4 portrait;margin:0;@top-left{{content:none;}}@top-right{{content:none;}}@bottom-left{{content:none;}}@bottom-right{{content:none;}}}}
"""

def build_html(centro, salidas, logo_path, fecha_entrega=None):
    if fecha_entrega is None: fecha_entrega=_hoy_es()
    logo_src=_logo_b64(logo_path) if logo_path else None
    logo_tag=(f'<img src="{logo_src}" alt="Alvale" style="height:25mm;width:auto;display:block;"/>' if logo_src else '<div style="height:25mm;width:50mm;background:#EEEFF7;border-radius:4px;"></div>')
    periodo=centro.get('año',''); nombre=_e(centro.get('razon_social','')); num_aut=_e(centro.get('num_aut',''))
    tipo_res='Gestión de Residuos de Vehículos al Final de su Vida Útil'
    css=_get_css(nombre=nombre,num_aut=num_aut,periodo=periodo)
    portada=f'''<section class="sheet--cover"><div style="flex:1;display:flex;flex-direction:column;padding:28mm 22mm 0;"><div style="flex:0.4;"></div><h1 style="font-size:34pt;font-weight:700;color:#1B1D58;line-height:0.95;text-transform:uppercase;letter-spacing:-0.01em;margin:0 0 4mm;font-family:'IBM Plex Sans',system-ui,sans-serif;">Registros<br/>Residuos Salida</h1><div style="font-size:9.5pt;font-weight:700;color:#2C2E88;text-transform:uppercase;letter-spacing:0.06em;margin-bottom:14mm;font-family:'IBM Plex Sans',system-ui,sans-serif;">{_e(tipo_res)}</div><div style="border:1px solid #E8E8EF;border-left:3px solid #2C2E88;border-radius:3px;overflow:hidden;"><div style="padding:3mm 4mm;background:#EEEFF7;font-size:9pt;font-weight:700;color:#2C2E88;letter-spacing:0.08em;text-transform:uppercase;font-family:'IBM Plex Sans',system-ui,sans-serif;">Datos del centro</div><table style="width:100%;border-collapse:collapse;font-family:'IBM Plex Sans',system-ui,sans-serif;"><tbody>{_cover_row('Entidad',nombre,True)}{_cover_row('NIMA',_e(centro.get('nima','')))}{_cover_row('Planta',_e(centro.get('direccion','')))}{_cover_row('Nº Autorización',num_aut)}{_cover_row('Tipo Autorización',_e(centro.get('tipo_aut','')))}{_cover_row('Periodo',_e(periodo),last=True)}</tbody></table></div><div style="flex:1;"></div><div style="text-align:right;font-size:9pt;color:#7E8094;padding-bottom:6mm;margin-top:8mm;font-family:'IBM Plex Sans',system-ui,sans-serif;">Fecha: {_e(fecha_entrega)}</div></div><div style="margin:0 22mm;border-top:0.25mm solid #D2D2DE;"></div><div style="padding:8mm 22mm 14mm;display:flex;justify-content:space-between;align-items:center;"><div style="display:flex;align-items:center;gap:6mm;">{logo_tag}<div style="font-size:8.5pt;color:#2A2C3D;line-height:1.2;font-family:'Poppins',system-ui,sans-serif;"><b style="color:#2C2E88;font-weight:700;font-size:9pt;display:block;margin-bottom:0.5mm;">Alvale Consulting Ingenieros, S.L.</b><span style="display:block;font-family:'IBM Plex Sans',system-ui,sans-serif;">Ribera de Axpe, 11, L-311-B. 48950 ERANDIO (Bizkaia)</span><span style="display:block;font-family:'IBM Plex Sans',system-ui,sans-serif;">Tl.: 944 971 050</span></div></div><div style="font-size:9pt;text-align:right;line-height:1.8;font-family:'Poppins',system-ui,sans-serif;"><a href="https://alvaleconsulting.com" style="color:#2C2E88;text-decoration:none;font-weight:600;display:block;">alvaleconsulting.com</a><a href="https://www.linkedin.com/company/alvale-consulting-ingenieros/" style="color:#2A2C3D;text-decoration:none;display:inline-flex;align-items:center;gap:1.5mm;"><span style="display:inline-flex;align-items:center;justify-content:center;width:5mm;height:5mm;border-radius:0.6mm;background:#2C2E88;color:#fff;font-weight:800;font-size:7.5pt;">in</span>LinkedIn</a></div></div></section>'''
    body_rows=''.join(_row(r) for r in salidas) or '<tr><td colspan="18" style="text-align:center;color:#7E8094;padding:6mm;font-style:italic;">No se han registrado movimientos.</td></tr>'
    sal_html=f'<section class="crono-section"><div class="doc-section"><h2 class="doc-section__title">Residuos retirados <span style="font-weight:400;color:#5D549D;">(salidas)</span></h2></div><table class="crono-table">{_colgroup()}{_thead()}<tbody>{body_rows}</tbody></table></section>'
    return f'<!doctype html>\n<html lang="es">\n<head>\n<meta charset="utf-8"/>\n<title>Archivo Cronológico — {nombre} · {periodo} · Alvale Consulting</title>\n<style>{css}</style>\n</head>\n<body>\n<div class="print-bar"><button onclick="window.print()">🖨️ Imprimir / Guardar como PDF</button></div>\n{portada}\n{sal_html}\n</body>\n</html>'

# ── Función principal ─────────────────────────────────────────────
def generar(excel_path, logo_path, output_path, fecha_entrega=None):
    import openpyxl
    wb=openpyxl.load_workbook(excel_path,data_only=True)
    centro=leer_datos_centro(wb); salidas=leer_salidas(wb)
    html=build_html(centro=centro,salidas=salidas,logo_path=logo_path,fecha_entrega=fecha_entrega)
    with open(output_path,'w',encoding='utf-8') as f: f.write(html)
    return output_path

# ── GUI ───────────────────────────────────────────────────────────
def lanzar_gui():
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    BLUE='#2C2E88'; PURP='#5D549D'; WHITE='#ffffff'; LIGHT='#EEEFF7'
    root=tk.Tk(); root.title('Archivo Cronológico VFU — Alvale Consulting')
    root.configure(bg=WHITE); root.resizable(False,False)
    hdr=tk.Frame(root,bg=BLUE,pady=12); hdr.pack(fill='x')
    tk.Label(hdr,text='GENERADOR — ARCHIVO CRONOLÓGICO VFU',font=('Helvetica',13,'bold'),bg=BLUE,fg=WHITE).pack()
    tk.Label(hdr,text='Alvale Consulting Ingenieros, S.L.',font=('Helvetica',9),bg=BLUE,fg='#c0c4e8').pack()
    body=tk.Frame(root,bg=WHITE,padx=22,pady=14); body.pack(fill='both')
    def section(title):
        f=tk.LabelFrame(body,text=f'  {title}  ',font=('Helvetica',9,'bold'),bg=WHITE,fg=PURP,bd=1,relief='groove',padx=10,pady=8)
        f.pack(fill='x',pady=(0,8)); return f
    def labeled_entry(parent,label,var,row,width=36):
        tk.Label(parent,text=label,font=('Helvetica',9),bg=WHITE,fg='#333',anchor='w',width=20).grid(row=row,column=0,sticky='w',pady=2)
        e=tk.Entry(parent,textvariable=var,width=width,font=('Helvetica',9),relief='solid',bd=1); e.grid(row=row,column=1,sticky='w',padx=(4,0),pady=2); return e
    def file_row(parent,label,var,row,exts):
        tk.Label(parent,text=label,font=('Helvetica',9),bg=WHITE,fg='#333',anchor='w',width=20).grid(row=row,column=0,sticky='w',pady=2)
        tk.Entry(parent,textvariable=var,width=36,font=('Helvetica',9),relief='solid',bd=1).grid(row=row,column=1,sticky='w',padx=(4,4),pady=2)
        tk.Button(parent,text='Examinar…',font=('Helvetica',8),bg=LIGHT,relief='flat',bd=1,cursor='hand2',command=lambda:var.set(filedialog.askopenfilename(filetypes=exts) or var.get())).grid(row=row,column=2,sticky='w')
    def save_row(parent,var,row):
        tk.Label(parent,text='HTML de salida:',font=('Helvetica',9),bg=WHITE,fg='#333',anchor='w',width=20).grid(row=row,column=0,sticky='w',pady=2)
        tk.Entry(parent,textvariable=var,width=36,font=('Helvetica',9),relief='solid',bd=1).grid(row=row,column=1,sticky='w',padx=(4,4),pady=2)
        tk.Button(parent,text='Guardar como…',font=('Helvetica',8),bg=LIGHT,relief='flat',bd=1,cursor='hand2',command=lambda:var.set(filedialog.asksaveasfilename(defaultextension='.html',filetypes=[('HTML','*.html'),('Todos','*')]) or var.get())).grid(row=row,column=2,sticky='w')
    sf=section('📂  Archivos')
    v_excel=tk.StringVar(); v_logo=tk.StringVar(value=LOGO_DEFAULT); v_out=tk.StringVar()
    file_row(sf,'Excel Ingurunet:',v_excel,0,[('Excel','*.xlsx'),('Todos','*')])
    file_row(sf,'Logo Alvale (.png):',v_logo,1,[('PNG','*.png'),('Todos','*')])
    save_row(sf,v_out,2)
    opt=section('📅  Opciones'); v_fecha=tk.StringVar(value=_hoy_es())
    labeled_entry(opt,'Fecha entrega:',v_fecha,0)
    bot=tk.Frame(body,bg=WHITE); bot.pack(fill='x',pady=(4,0))
    pb=ttk.Progressbar(bot,mode='indeterminate',length=420); pb.pack(side='left',padx=(0,10))
    lbl=tk.Label(bot,text='',font=('Helvetica',8),bg=WHITE,fg=PURP); lbl.pack(side='left')
    def on_generar():
        missing=[f for f,v in [('Excel Ingurunet',v_excel),('HTML de salida',v_out)] if not v.get()]
        if missing: messagebox.showerror('Faltan datos','Completa:\n• '+'\n• '.join(missing)); return
        btn.config(state='disabled'); lbl.config(text='Generando HTML…',fg=PURP); pb.start(12); root.update()
        def worker():
            try:
                out=generar(excel_path=v_excel.get(),logo_path=v_logo.get(),output_path=v_out.get(),fecha_entrega=v_fecha.get() or None)
                pb.stop(); lbl.config(text='✔  HTML generado correctamente',fg='#1a7a3a'); btn.config(state='normal')
                if messagebox.askyesno('¡Listo!',f'HTML generado:\n{out}\n\n¿Abrir en el navegador?'):
                    webbrowser.open('file:///'+os.path.abspath(out).replace('\\','/'))
            except Exception as ex:
                pb.stop(); lbl.config(text='✖  Error al generar',fg='#c0392b'); btn.config(state='normal')
                messagebox.showerror('Error',str(ex))
        threading.Thread(target=worker,daemon=True).start()
    btn=tk.Button(body,text='  ▶  GENERAR ARCHIVO CRONOLÓGICO  ',font=('Helvetica',11,'bold'),bg=BLUE,fg=WHITE,activebackground=PURP,activeforeground=WHITE,relief='flat',cursor='hand2',pady=9,command=on_generar)
    btn.pack(fill='x',pady=(6,14))
    root.mainloop()

if __name__=='__main__':
    lanzar_gui()
